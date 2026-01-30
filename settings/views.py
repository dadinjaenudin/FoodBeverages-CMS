"""
Settings Views - Bulk Import Excel
"""
import logging
import traceback
import uuid
import shutil
import os
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from products.models import Category, Product, Modifier, ModifierOption, ProductPhoto, ProductModifier
from core.models import Brand, Company

# Import log saving function
from .import_log_views import save_import_log

logger = logging.getLogger(__name__)


@login_required
def bulk_delete_products_view(request):
    """Display bulk delete products page"""
    # Get brands for selection
    brands = Brand.objects.select_related('company').filter(is_active=True).order_by('company__name', 'name')
    
    # Count data per brand for display
    brand_data = []
    for brand in brands:
        brand_data.append({
            'brand': brand,
            'product_count': Product.objects.filter(brand=brand).count(),
            'category_count': Category.objects.filter(brand=brand).count(),
            'modifier_count': Modifier.objects.filter(brand=brand).count(),
        })
    
    context = {
        'brands': brands,
        'brand_data': brand_data,
    }
    return render(request, 'settings/bulk_delete_products.html', context)


@login_required
@require_http_methods(["POST"])
def bulk_delete_products_action(request):
    """
    Bulk delete products, categories, and modifiers by brand
    DANGER: This is irreversible!
    """
    try:
        brand_id = request.POST.get('brand_id')
        delete_products = request.POST.get('delete_products') == 'on'
        delete_categories = request.POST.get('delete_categories') == 'on'
        delete_modifiers = request.POST.get('delete_modifiers') == 'on'
        confirm_delete = request.POST.get('confirm_delete') == 'on'
        
        if not brand_id or not confirm_delete:
            return JsonResponse({
                'success': False,
                'message': 'Brand selection and confirmation are required'
            }, status=400)
        
        if not any([delete_products, delete_categories, delete_modifiers]):
            return JsonResponse({
                'success': False,
                'message': 'Please select at least one data type to delete'
            }, status=400)
        
        # Get brand
        try:
            brand = Brand.objects.get(id=brand_id)
        except Brand.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Brand not found'
            }, status=404)
        
        stats = {
            'products_deleted': 0,
            'categories_deleted': 0,
            'modifiers_deleted': 0,
            'errors': []
        }
        
        with transaction.atomic():
            # Delete products first (due to foreign key constraints)
            if delete_products:
                products = Product.objects.filter(brand=brand)
                product_count = products.count()
                
                # Delete product photos first
                for product in products:
                    # Delete photo files
                    for photo in product.photos.all():
                        try:
                            if photo.photo and os.path.exists(photo.photo.path):
                                os.remove(photo.photo.path)
                        except Exception as e:
                            stats['errors'].append(f"Could not delete photo {photo.photo}: {str(e)}")
                
                # Delete product photos from DB
                ProductPhoto.objects.filter(product__brand=brand).delete()
                
                # Delete product-modifier relationships
                ProductModifier.objects.filter(product__brand=brand).delete()
                
                # Delete products
                products.delete()
                stats['products_deleted'] = product_count
                logger.info(f"Deleted {product_count} products for brand {brand.name}")
            
            # Delete modifiers
            if delete_modifiers:
                # Delete modifier options first
                modifier_options = ModifierOption.objects.filter(modifier__brand=brand)
                modifier_option_count = modifier_options.count()
                modifier_options.delete()
                
                # Delete modifiers
                modifiers = Modifier.objects.filter(brand=brand)
                modifier_count = modifiers.count()
                modifiers.delete()
                
                stats['modifiers_deleted'] = modifier_count + modifier_option_count
                logger.info(f"Deleted {modifier_count} modifiers and {modifier_option_count} options for brand {brand.name}")
            
            # Delete categories
            if delete_categories:
                categories = Category.objects.filter(brand=brand)
                category_count = categories.count()
                categories.delete()
                stats['categories_deleted'] = category_count
                logger.info(f"Deleted {category_count} categories for brand {brand.name}")
        
        logger.info(f"Bulk delete completed by {request.user.username} for brand {brand.name}: {stats}")
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted data for brand: {brand.name}',
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Bulk delete error: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error during deletion: {str(e)}'
        }, status=500)


@login_required
def bulk_import_two_sheet_view(request):
    """Display bulk import two sheet page"""
    return render(request, 'settings/bulk_import_two_sheet.html')


@login_required
@require_http_methods(["POST"])
def upload_two_sheet_excel(request):
    """
    Upload and process two-sheet Excel format
    Process Groups FIRST, then Products (two-file system logic in one file)
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)
        
        excel_file = request.FILES['file']
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Please upload .xlsx or .xls file'
            }, status=400)
        
        # Load workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Get import options
        skip_duplicates = request.POST.get('skip_duplicates') == 'on'
        update_existing = request.POST.get('update_existing') == 'on'
        create_modifiers = request.POST.get('create_modifiers') == 'on'
        
        # Update option overrides skip option
        if update_existing:
            skip_duplicates = False
        
        stats = {
            'categories_created': 0,
            'products_created': 0,
            'products_updated': 0,
            'products_skipped': 0,
            'modifiers_created': 0,
            'modifier_options_created': 0,
            'product_modifiers_created': 0,
            'companies_created': 0,
            'brands_created': 0,
            'errors': [],
            'processed_rows': 0,
            'successful_rows': 0,
            'failed_rows': [],
            'detailed_log': []
        }
        
        # Define static image folder
        static_image_folder = Path(settings.BASE_DIR) / 'static' / 'product' / 'image'
        logger.info(f"DEBUG: Looking for images in: {static_image_folder.absolute()}")
        if not static_image_folder.exists():
            logger.warning(f"DEBUG: Static image folder does not exist! Creating it.")
        
        # FIX: Use product_images to match Product.image field
        media_product_folder = Path(settings.MEDIA_ROOT) / 'product_images'
        
        # Create folders if not exist
        static_image_folder.mkdir(parents=True, exist_ok=True)
        media_product_folder.mkdir(parents=True, exist_ok=True)
        
        # Cache for brands and categories
        brand_cache = {}
        category_cache = {}
        modifier_cache = {}
        
        # Get default brand for condiment groups (since they don't have company/brand codes)
        default_brand = None
        try:
            default_brand = Brand.objects.first()
            if not default_brand:
                return JsonResponse({
                    'success': False,
                    'message': 'No brand found in system. Please create a brand first.'
                }, status=400)
            logger.info(f"Using default brand for condiment groups: {default_brand.name} (ID: {default_brand.id})")
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error getting default brand: {str(e)}'
            }, status=400)
        
        with transaction.atomic():
            # ===== STEP 1: PROCESS CONDIMENT GROUPS FIRST =====
            if 'Condiment Groups' in wb.sheetnames and create_modifiers:
                ws_condiments = wb['Condiment Groups']
                logger.info("Processing Condiment Groups sheet FIRST...")
                
                # Get header row
                headers = [cell.value for cell in ws_condiments[1]]
                col_indices = {}
                for idx, header in enumerate(headers):
                    if header:
                        col_indices[header.strip()] = idx
                
                # Validate required columns
                required_cols = ['Group Name', 'Option Name', 'Fee']
                missing_cols = [col for col in required_cols if col not in col_indices]
                if missing_cols:
                    stats['errors'].append(f'Missing required columns in Condiment Groups: {", ".join(missing_cols)}')
                else:
                    # Skip header row
                    for row_idx, row in enumerate(ws_condiments.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            group_name = row[col_indices['Group Name']]
                            option_name = row[col_indices['Option Name']]
                            fee = row[col_indices['Fee']]
                            is_required = row[col_indices.get('Is Required', -1)] if 'Is Required' in col_indices else 'No'
                            max_selections = row[col_indices.get('Max Selections', -1)] if 'Max Selections' in col_indices else 1
                            
                            if not group_name or not option_name:
                                continue
                            
                            group_name = str(group_name).strip()
                            option_name = str(option_name).strip()
                            
                            # Process Is Required
                            is_req = False
                            if is_required:
                                is_req_str = str(is_required).strip().lower()
                                is_req = is_req_str in ['yes', 'y', 'true', '1']
                            
                            # Process Max Selections
                            try:
                                max_sel = int(max_selections) if max_selections else 1
                            except:
                                max_sel = 1
                            
                            # Create or get modifier group
                            if group_name not in modifier_cache:
                                modifier, created = Modifier.objects.get_or_create(
                                    name=group_name,
                                    brand=default_brand,  # Use default_brand here
                                    defaults={
                                        'is_required': is_req,
                                        'max_selections': max_sel,
                                        'is_active': True
                                    }
                                )
                                modifier_cache[group_name] = modifier
                                if created:
                                    stats['modifiers_created'] += 1
                            else:
                                modifier = modifier_cache[group_name]
                            
                            # Create modifier option
                            option, opt_created = ModifierOption.objects.get_or_create(
                                modifier=modifier,
                                name=option_name,
                                defaults={
                                    'price_adjustment': Decimal(str(fee)) if fee else Decimal('0'),
                                    'is_active': True
                                }
                            )
                            
                            if opt_created:
                                stats['modifier_options_created'] += 1
                            elif update_existing:
                                new_fee = Decimal(str(fee)) if fee else Decimal('0')
                                if option.price_adjustment != new_fee:
                                    option.price_adjustment = new_fee
                                    option.save()
                            
                        except Exception as e:
                            stats['errors'].append(f"Row {row_idx} (Condiment Groups): {str(e)}")
                            logger.error(f"Error processing condiment row {row_idx}: {traceback.format_exc()}")
            
            # ===== STEP 2: PROCESS PRODUCTS SECOND =====
            if 'Products' in wb.sheetnames:
                ws_products = wb['Products']
                logger.info("Processing Products sheet SECOND...")
                
                # Get header row
                headers = [cell.value for cell in ws_products[1]]
                col_indices = {}
                for idx, header in enumerate(headers):
                    if header:
                        col_indices[header.strip()] = idx
                
                # Validate required columns (flexible naming)
                required_cols_mapping = {
                    'Company Code': ['Company Code'],
                    'Brand Code': ['Brand Code'], 
                    'Category': ['Category'],
                    'Menu Category': ['Menu Category'],
                    'Product Name': ['Product Name', 'Nama Product', 'Product_Name'],
                    'PLU Product': ['PLU Product', 'PLU_Product'],
                    'Price Product': ['Price Product', 'Price_Product']
                }
                
                # Check if all required columns exist (with flexible naming)
                missing_cols = []
                for required_col, alternatives in required_cols_mapping.items():
                    if not any(alt in col_indices for alt in alternatives):
                        missing_cols.append(f"{required_col} (tried: {', '.join(alternatives)})")
                
                if missing_cols:
                    stats['errors'].append(f'Missing required columns in Products: {", ".join(missing_cols)}')
                else:
                    # Skip header row
                    for row_idx, row in enumerate(ws_products.iter_rows(min_row=2, values_only=True), start=2):
                        stats['processed_rows'] += 1
                        
                        try:
                            # Get data using flexible column names
                            def get_col_value(col_mappings):
                                for col_name in col_mappings:
                                    if col_name in col_indices:
                                        return row[col_indices[col_name]]
                                return None
                            
                            company_code = get_col_value(['Company Code'])
                            brand_code = get_col_value(['Brand Code'])
                            category_name = get_col_value(['Category'])
                            menu_category_name = get_col_value(['Menu Category'])
                            product_name = get_col_value(['Product Name', 'Nama Product', 'Product_Name'])
                            plu_product = str(get_col_value(['PLU Product', 'PLU_Product'])) if get_col_value(['PLU Product', 'PLU_Product']) else ''
                            price_product = get_col_value(['Price Product', 'Price_Product'])
                            
                            # Log raw data
                            log_entry = {
                                'row': row_idx,
                                'raw_data': {
                                    'company_code': company_code,
                                    'brand_code': brand_code,
                                    'category': category_name,
                                    'menu_category': menu_category_name,
                                    'product_name': product_name,
                                    'plu_product': plu_product,
                                    'price': price_product
                                },
                                'status': 'processing',
                                'steps': []
                            }
                            
                            # Optional columns
                            printer_kitchen = get_col_value(['Printer Kitchen']) if 'Printer Kitchen' in col_indices else 'kitchen'
                            condiment_groups = get_col_value(['Condiment Groups']) if 'Condiment Groups' in col_indices else None
                            # FIX: Support multiple column names for Image Product
                            image_product = get_col_value(['Image Product', 'Image', 'Product Image', 'Image_Product'])
                            
                            if not company_code or not brand_code or not product_name or not plu_product:
                                log_entry['status'] = 'skipped'
                                log_entry['reason'] = 'Missing required fields'
                                log_entry['missing_fields'] = []
                                if not company_code:
                                    log_entry['missing_fields'].append('company_code')
                                if not brand_code:
                                    log_entry['missing_fields'].append('brand_code')
                                if not product_name:
                                    log_entry['missing_fields'].append('product_name')
                                if not plu_product:
                                    log_entry['missing_fields'].append('plu_product')
                                stats['detailed_log'].append(log_entry)
                                continue
                            
                            # ===== GET OR CREATE BRAND BY CODE =====
                            brand_key = f"{company_code.strip()}_{brand_code.strip()}"
                            log_entry['steps'].append(f'Processing brand: {brand_key}')
                            
                            if brand_key not in brand_cache:
                                try:
                                    brand = Brand.objects.select_related('company').get(
                                        code=brand_code.strip(),
                                        company__code=company_code.strip()
                                    )
                                    brand_cache[brand_key] = brand
                                    log_entry['steps'].append(f'Found existing brand: {brand.name} (ID: {brand.id})')
                                except Brand.DoesNotExist:
                                    # Auto-create company if it doesn't exist
                                    log_entry['steps'].append('Brand not found, creating company...')
                                    try:
                                        company = Company.objects.get(code=company_code.strip())
                                        log_entry['steps'].append(f'Found existing company: {company.name} (ID: {company.id})')
                                    except Company.DoesNotExist:
                                        # Create new company with auto-generated name
                                        company_name = f"Company {company_code.strip()}"
                                        company = Company.objects.create(
                                            code=company_code.strip(),
                                            name=company_name,
                                            is_active=True
                                        )
                                        stats['companies_created'] += 1
                                        log_entry['steps'].append(f'Created new company: {company.name} (ID: {company.id})')
                                    
                                    # Create new brand with auto-generated name
                                    log_entry['steps'].append('Creating new brand...')
                                    brand_name = f"Brand {brand_code.strip()}"
                                    brand = Brand.objects.create(
                                        code=brand_code.strip(),
                                        name=brand_name,
                                        company=company,
                                        is_active=True
                                    )
                                    brand_cache[brand_key] = brand
                                    stats['brands_created'] += 1
                                    log_entry['steps'].append(f'Created new brand: {brand.name} (ID: {brand.id})')
                            else:
                                brand = brand_cache[brand_key]
                                log_entry['steps'].append(f'Using cached brand: {brand.name}')
                            
                            # Auto-create category if empty
                            if not category_name:
                                category_name = "Uncategorized"
                            
                            # ===== PROCESS CATEGORIES =====
                            parent_category = None
                            final_category = None
                            
                            # Create/get parent category (use category_name if menu_category is empty)
                            if category_name:
                                cat_key = f"{brand.id}_{category_name.strip()}"
                                if cat_key not in category_cache:
                                    parent_category, created = Category.objects.get_or_create(
                                        brand=brand,
                                        name=category_name.strip(),
                                        parent=None,
                                        defaults={'sort_order': 0, 'is_active': True}
                                    )
                                    category_cache[cat_key] = parent_category
                                    if created:
                                        stats['categories_created'] += 1
                                        logger.info(f"Created parent category: {category_name}")
                                else:
                                    parent_category = category_cache[cat_key]
                            
                            # Create/get menu category (child) - use menu_category or fallback to category
                            child_category_name = menu_category_name if menu_category_name else f"{category_name} Items"
                            if child_category_name:
                                menu_cat_key = f"{brand.id}_{child_category_name.strip()}"
                                if menu_cat_key not in category_cache:
                                    final_category, created = Category.objects.get_or_create(
                                        brand=brand,
                                        name=child_category_name.strip(),
                                        parent=parent_category,
                                        defaults={'sort_order': 0, 'is_active': True}
                                    )
                                    category_cache[menu_cat_key] = final_category
                                    if created:
                                        stats['categories_created'] += 1
                                        logger.info(f"Created child category: {child_category_name}")
                                else:
                                    final_category = category_cache[menu_cat_key]
                            
                            # Fallback: if no categories at all, use parent as final
                            if not final_category:
                                final_category = parent_category
                            
                            # Last resort: create default category
                            if not final_category:
                                default_cat_key = f"{brand.id}_Default"
                                if default_cat_key not in category_cache:
                                    final_category, created = Category.objects.get_or_create(
                                        brand=brand,
                                        name="Default",
                                        parent=None,
                                        defaults={'sort_order': 0, 'is_active': True}
                                    )
                                    category_cache[default_cat_key] = final_category
                                    if created:
                                        stats['categories_created'] += 1
                                        logger.info("Created default category")
                                else:
                                    final_category = category_cache[default_cat_key]
                            
                            # ===== PROCESS PRODUCT =====
                            # Check for duplicate product using business logic (full combination)
                            existing_product = None
                            if skip_duplicates or update_existing:
                                try:
                                    existing_product = Product.objects.get(
                                        brand=brand,           # ← Brand Code (via brand object)
                                        category=final_category,  # ← Category + Menu Category (via category object)
                                        name=product_name.strip(),  # ← Product Name
                                        sku=plu_product,          # ← PLU Product
                                    )
                                except Product.DoesNotExist:
                                    existing_product = None
                            
                            # Skip if duplicate and not updating
                            if skip_duplicates and existing_product:
                                log_entry['status'] = 'skipped'
                                log_entry['reason'] = 'Exact duplicate product found (Business Logic)'
                                log_entry['duplicate_key'] = {
                                    'company_code': company_code,  # ← Added Company Code
                                    'brand_code': brand_code,     # ← Added Brand Code
                                    'category': category_name,
                                    'menu_category': menu_category_name,
                                    'product_name': product_name,
                                    'plu_product': plu_product,
                                    'existing_product_id': str(existing_product.id)
                                }
                                stats['products_skipped'] += 1
                                stats['detailed_log'].append(log_entry)
                                logger.info(f"Skipped exact duplicate: {product_name} (Company: {company_code}, Brand: {brand_code}, Category: {final_category.name}, SKU: {plu_product})")
                                continue
                            
                            # Map printer_kitchen
                            printer_map = {
                                'bar': 'bar',
                                'kitchen': 'kitchen',
                                'dessert': 'dessert'
                            }
                            printer_target = printer_map.get(str(printer_kitchen).strip().lower(), 'kitchen')
                            
                            # Create or update product
                            if update_existing and existing_product:
                                # Update existing product
                                log_entry['steps'].append('Updating existing product...')
                                existing_product.name = product_name.strip()
                                existing_product.company = brand.company  # FIX: Ensure company is set
                                existing_product.category = final_category
                                existing_product.description = ''
                                existing_product.price = Decimal(str(price_product)) if price_product else Decimal('0')
                                existing_product.printer_target = printer_target
                                existing_product.save()
                                
                                product = existing_product
                                stats['products_updated'] += 1
                                log_entry['status'] = 'updated'
                                log_entry['product_id'] = str(product.id)
                                log_entry['steps'].append(f'Updated product: {product.name}')
                            else:
                                # Create new product using business logic
                                log_entry['steps'].append('Creating new product...')
                                
                                # Use get_or_create to avoid transaction issues
                                product, created = Product.objects.get_or_create(
                                    brand=brand,           # ← Brand Code (via brand object)
                                    category=final_category,  # ← Category + Menu Category (via category object)
                                    name=product_name.strip(),  # ← Product Name
                                    sku=plu_product,          # ← PLU Product
                                    defaults={
                                        'company': brand.company,  # FIX: Ensure company is set
                                        'description': '',
                                        'price': Decimal(str(price_product)) if price_product else Decimal('0'),
                                        'cost': Decimal('0'),
                                        'printer_target': printer_target,
                                        'track_stock': False,
                                        'stock_quantity': Decimal('0'),
                                        'sort_order': 0,
                                        'is_active': True
                                    }
                                )
                                
                                if created:
                                    stats['products_created'] += 1
                                    log_entry['status'] = 'created'
                                    log_entry['product_id'] = str(product.id)
                                    log_entry['steps'].append(f'Created new product: {product.name} (ID: {product.id})')
                                else:
                                    stats['products_updated'] += 1
                                    log_entry['status'] = 'updated'
                                    log_entry['product_id'] = str(product.id)
                                    log_entry['steps'].append(f'Product already existed with same business logic')
                                    
                                    # Update existing product if data differs
                                    needs_update = False
                                    if product.name != product_name.strip():
                                        product.name = product_name.strip()
                                        needs_update = True
                                    if product.company != brand.company:
                                        product.company = brand.company
                                        needs_update = True
                                    if product.category != final_category:
                                        product.category = final_category
                                        needs_update = True
                                    if price_product and Decimal(str(price_product)) != product.price:
                                        product.price = Decimal(str(price_product))
                                        needs_update = True
                                    
                                    if needs_update:
                                        product.save()
                                        log_entry['steps'].append('Updated existing product data')
                            
                            stats['successful_rows'] += 1
                            stats['detailed_log'].append(log_entry)
                            
                            # ===== PROCESS IMAGE =====
                            if image_product and str(image_product).strip():
                                try:
                                    # Extract filename only (remove path)
                                    image_path = str(image_product).strip()
                                    filename = Path(image_path).name  # This gets only the filename
                                    
                                    # If no extension, try to add common ones
                                    if '.' not in filename:
                                        possible_extensions = ['.jpg', '.jpeg', '.png', '.webp']
                                        for ext in possible_extensions:
                                            test_filename = filename + ext
                                            test_path = static_image_folder / test_filename
                                            if test_path.exists():
                                                filename = test_filename
                                                break
                                    
                                    source_path = static_image_folder / filename
                                    
                                    if source_path.exists():
                                        # Use ORIGINAL filename instead of UUID (User Request)
                                        dest_path = media_product_folder / filename
                                        
                                        # Copy file to media folder (overwrite if exists)
                                        shutil.copy2(source_path, dest_path)
                                        
                                        # Save to Product.image
                                        product.image = f'product_images/{filename}'
                                        product.save()

                                        # Create ProductPhoto as well for backward compatibility / gallery
                                        # Use get_or_create to avoid duplicates if using same filename
                                        ProductPhoto.objects.get_or_create(
                                            product=product,
                                            photo=f'product_images/{filename}', 
                                            defaults={'is_primary': True, 'sort_order': 0}
                                        )
                                        logger.info(f"Image processed: {filename}")
                                    
                                    # FALLBACK: Check if image already exists in media folder
                                    elif (media_product_folder / filename).exists():
                                        logger.info(f"Image found in media folder: {filename}")
                                        product.image = f'product_images/{filename}'
                                        product.save()
                                        
                                        ProductPhoto.objects.get_or_create(
                                            product=product,
                                            photo=f'product_images/{filename}', 
                                            defaults={'is_primary': True, 'sort_order': 0}
                                        )
                                    else:
                                        logger.error(f"DEBUG: Image NOT FOUND at: {source_path.absolute()}")
                                        stats['errors'].append(
                                            f"Row {row_idx}: Image '{filename}' not found in static/product/image/ or media/product_images/"
                                        )
                                except Exception as img_error:
                                    stats['errors'].append(
                                        f"Row {row_idx}: Error processing image: {str(img_error)}"
                                    )
                            
                            # ===== LINK PRODUCTS TO MODIFIERS =====
                            if condiment_groups and str(condiment_groups).strip():
                                groups_str = str(condiment_groups).strip()
                                if groups_str and groups_str not in ['', 'nan', 'None']:
                                    group_names = [g.strip() for g in groups_str.split(',') if g.strip()]
                                    for group_name in group_names:
                                        if group_name in modifier_cache:
                                            modifier = modifier_cache[group_name]
                                            # Link product to modifier
                                            product_modifier, created = ProductModifier.objects.get_or_create(
                                                product=product,
                                                modifier=modifier,
                                                defaults={'sort_order': len(product.product_modifiers.all())}
                                            )
                                            if created:
                                                stats['product_modifiers_created'] += 1
                                        else:
                                            stats['errors'].append(
                                                f"Row {row_idx}: Condiment group '{group_name}' not found. Make sure it's defined in Condiment Groups sheet."
                                            )
                        
                        except Exception as e:
                            error_msg = f"Row {row_idx} (Products): {str(e)}"
                            stats['errors'].append(error_msg)
                            
                            # Add to failed rows with detailed error
                            log_entry['status'] = 'error'
                            log_entry['error'] = str(e)
                            log_entry['error_type'] = type(e).__name__
                            stats['failed_rows'].append({
                                'row': row_idx,
                                'product_name': product_name if 'product_name' in locals() else 'Unknown',
                                'error': str(e),
                                'error_type': type(e).__name__
                            })
                            stats['detailed_log'].append(log_entry)
                            logger.error(f"Error processing product row {row_idx}: {traceback.format_exc()}")
        
        logger.info(f"Two-sheet import completed by {request.user.username}: {stats}")
        
        # Save import log for debugging
        save_import_log(request, stats)
        
        return JsonResponse({
            'success': True,
            'message': 'Two-sheet file uploaded and processed successfully',
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Two-sheet import error: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        }, status=500)


@login_required
def download_two_sheet_template(request):
    """
    Generate and download Excel template with 2 sheets:
    - Products (simplified structure)
    - Condiment Groups (simplified structure without Company/Brand codes)
    """
    wb = Workbook()
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # ===== SHEET 1: PRODUCTS =====
    ws_products = wb.active
    ws_products.title = "Products"
    
    # Products headers - dengan Company dan Brand codes
    prod_headers = [
        'Company Code', 'Brand Code', 'Category', 'Menu Category', 'Product Name', 
        'PLU Product', 'Printer Kitchen', 'Condiment Groups', 'Price Product', 'Image Product'
    ]
    for col, header in enumerate(prod_headers, start=1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data dengan Company dan Brand codes
    ws_products.append([
        'YGY', 'AVRIL', 'Makanan', 'Ayam', 'Ayam Geprek Original',
        'AYG-001', 'kitchen', 'Tingkat Kepedasan,Ukuran', 25000, 'ayam-geprek-001.jpg'
    ])
    ws_products.append([
        'YGY', 'AVRIL', 'Minuman', 'Kopi', 'Cappuccino',
        'KCP-001', 'bar', 'Ukuran,Sugar Level,Temperature', 18000, 'cappuccino-001.jpg'
    ])
    ws_products.append([
        'YGY', 'AVRIL', 'Minuman', 'Teh', 'Teh Manis Dingin',
        'TMD-001', 'bar', '', 8000, 'teh-dingin-001.jpg'
    ])
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D']:
        ws_products.column_dimensions[col].width = 20
    for col in ['E', 'F', 'G', 'H']:
        ws_products.column_dimensions[col].width = 15
    
    # ===== SHEET 2: CONDIMENT GROUPS =====
    ws_condiments = wb.create_sheet("Condiment Groups")
    
    # Condiment Groups headers - simplified structure (no Company/Brand codes)
    cond_headers = [
        'Group Name', 'Option Name', 'Fee', 'Is Required', 'Max Selections'
    ]
    for col, header in enumerate(cond_headers, start=1):
        cell = ws_condiments.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data for "Tingkat Kepedasan"
    ws_condiments.append(['Tingkat Kepedasan', 'Level 1 (Tidak Pedas)', 0, 'No', 1])
    ws_condiments.append(['Tingkat Kepedasan', 'Level 3 (Pedas)', 0, 'No', 1])
    ws_condiments.append(['Tingkat Kepedasan', 'Level 5 (Sangat Pedas)', 0, 'No', 1])
    
    # Example data for "Ukuran"
    ws_condiments.append(['Ukuran', 'Regular', 0, 'Yes', 1])
    ws_condiments.append(['Ukuran', 'Large', 5000, 'No', 1])
    
    # Example data for "Sugar Level"
    ws_condiments.append(['Sugar Level', 'Less Sugar', 0, 'No', 1])
    ws_condiments.append(['Sugar Level', 'Normal Sugar', 0, 'No', 1])
    ws_condiments.append(['Sugar Level', 'Extra Sugar', 0, 'No', 1])
    
    # Example data for "Temperature"
    ws_condiments.append(['Temperature', 'Hot', 0, 'No', 1])
    ws_condiments.append(['Temperature', 'Ice', 0, 'No', 1])
    
    # Set column widths
    ws_condiments.column_dimensions['A'].width = 20
    ws_condiments.column_dimensions['B'].width = 25
    ws_condiments.column_dimensions['C'].width = 12
    ws_condiments.column_dimensions['D'].width = 15
    ws_condiments.column_dimensions['E'].width = 18
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return as downloadable file
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=two_sheet_import_template.xlsx'
    
    logger.info(f"Two-sheet template downloaded by user: {request.user.username}")
    return response


@login_required
def download_products_template(request):
    """
    Generate and download Excel template for custom products import format
    Single sheet with 10 columns format
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Import Template"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers for custom format
    headers = [
        'Company Code', 'Brand Code', 'Category', 'Menu Category', 'Product Name', 
        'PLU Product', 'Printer_Kitchen', 'Condiment Groups', 'Price Product', 'Image Product'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data
    ws.append([
        'YGY', 'AVRIL', 'Makanan', 'Ayam', 'Ayam Geprek Original',
        'AYG-001', 'kitchen', 'Tingkat Kepedasan,Ukuran', 25000, 'ayam-geprek-001.jpg'
    ])
    ws.append([
        'YGY', 'AVRIL', 'Minuman', 'Kopi', 'Cappuccino',
        'KCP-001', 'bar', 'Ukuran,Sugar Level', 18000, 'cappuccino-001.jpg'
    ])
    ws.append([
        'YGY', 'AVRIL', 'Minuman', 'Teh', 'Teh Manis Dingin',
        'TMD-001', 'bar', '', 8000, 'teh-dingin-001.jpg'
    ])
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    for col in ['F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 15
    
    # Add notes
    ws.append([''])  # Empty row
    ws.append(['NOTES:'])
    ws.append(['• Company Code: Your company code (e.g., YGY)'])
    ws.append(['• Brand Code: Your brand code (e.g., AVRIL)'])
    ws.append(['• Category: Parent category (Makanan, Minuman, Dessert, etc)'])
    ws.append(['• Menu Category: Child category (Ayam, Kopi, Teh, etc)'])
    ws.append(['• PLU Product: Unique SKU code per brand'])
    ws.append(['• Printer_Kitchen: bar/kitchen/dessert'])
    ws.append(['• Condiment Groups: Modifier names separated by comma (optional)'])
    ws.append(['• Price Product: Selling price (numbers only)'])
    ws.append(['• Image Product: Image filename (place in static/product/image/)'])
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return as downloadable file
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products_import_template.xlsx'
    
    logger.info(f"Products template downloaded by user: {request.user.username}")
    return response


@login_required
def bulk_import_view(request):
    """Display bulk import page"""
    return render(request, 'settings/bulk_import.html')


@login_required
def download_template(request):
    """
    Generate and download Excel template with 3 sheets:
    - Categories
    - Products
    - Modifiers
    """
    wb = Workbook()
    
    # ===== SHEET 1: CATEGORIES =====
    ws_category = wb.active
    ws_category.title = "Categories"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Categories headers
    cat_headers = [
        'Company Code', 'Brand Code', 'Category Name', 'Parent Category', 'Sort Order', 'Icon', 'Is Active'
    ]
    for col, header in enumerate(cat_headers, start=1):
        cell = ws_category.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data
    ws_category.append(['YGY', 'AVRIL', 'Makanan', '', '1', 'fa-utensils', 'TRUE'])
    ws_category.append(['YGY', 'AVRIL', 'Minuman', '', '2', 'fa-glass', 'TRUE'])
    ws_category.append(['YGY', 'AVRIL', 'Ayam', 'Makanan', '3', '', 'TRUE'])
    
    # Set column widths
    ws_category.column_dimensions['A'].width = 15
    ws_category.column_dimensions['B'].width = 15
    ws_category.column_dimensions['C'].width = 25
    ws_category.column_dimensions['D'].width = 25
    ws_category.column_dimensions['E'].width = 12
    ws_category.column_dimensions['F'].width = 15
    ws_category.column_dimensions['G'].width = 12
    
    # ===== SHEET 2: PRODUCTS =====
    ws_product = wb.create_sheet("Products")
    
    prod_headers = [
        'Company Code', 'Brand Code', 'SKU', 'Product Name', 'Category Name', 'Description',
        'Price', 'Cost', 'Printer Target', 'Track Stock', 'Stock Quantity',
        'Image Filename', 'Sort Order', 'Is Active'
    ]
    for col, header in enumerate(prod_headers, start=1):
        cell = ws_product.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data
    ws_product.append([
        'YGY', 'AVRIL', 'AYG-001', 'Ayam Geprek Original', 'Ayam', 'Ayam geprek dengan sambal level 5',
        25000, 15000, 'kitchen', 'FALSE', 0, 
        'ayam-geprek-001.jpg', 1, 'TRUE'
    ])
    ws_product.append([
        'YGY', 'AVRIL', 'MJU-001', 'Jus Jeruk', 'Minuman', 'Jus jeruk segar',
        15000, 8000, 'bar', 'TRUE', 50,
        'jus-jeruk-001.jpg', 2, 'TRUE'
    ])
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_product.column_dimensions[col].width = 20
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws_product.column_dimensions[col].width = 15
    
    # ===== SHEET 3: MODIFIERS =====
    ws_modifier = wb.create_sheet("Modifiers")
    
    mod_headers = [
        'Company Code', 'Brand Code', 'Modifier Group', 'Is Required', 'Max Selections',
        'Option Name', 'Price Adjustment', 'Is Default', 'Sort Order', 'Is Active'
    ]
    for col, header in enumerate(mod_headers, start=1):
        cell = ws_modifier.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data
    ws_modifier.append([
        'YGY', 'AVRIL', 'Tingkat Kepedasan', 'TRUE', '1',
        'Level 1 (Tidak Pedas)', 0, 'FALSE', 1, 'TRUE'
    ])
    ws_modifier.append([
        'YGY', 'AVRIL', 'Tingkat Kepedasan', 'TRUE', '1',
        'Level 3 (Pedas)', 0, 'TRUE', 2, 'TRUE'
    ])
    ws_modifier.append([
        'YGY', 'AVRIL', 'Tingkat Kepedasan', 'TRUE', '1',
        'Level 5 (Sangat Pedas)', 0, 'FALSE', 3, 'TRUE'
    ])
    ws_modifier.append([
        'YGY', 'AVRIL', 'Ukuran', 'FALSE', '1',
        'Regular', 0, 'TRUE', 1, 'TRUE'
    ])
    ws_modifier.append([
        'YGY', 'AVRIL', 'Ukuran', 'FALSE', '1',
        'Large', 5000, 'FALSE', 2, 'TRUE'
    ])
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_modifier.column_dimensions[col].width = 20
    for col in ['F', 'G', 'H', 'I']:
        ws_modifier.column_dimensions[col].width = 15
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return as downloadable file
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_import_template.xlsx'
    
    logger.info(f"Template downloaded by user: {request.user.username}")
    return response


@login_required
@require_http_methods(["POST"])
def upload_excel(request):
    """
    Upload and process Excel file with bulk data
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)
        
        excel_file = request.FILES['file']
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Please upload .xlsx or .xls file'
            }, status=400)
        
        # Load workbook
        wb = load_workbook(excel_file, data_only=True)
        
        stats = {
            'categories_created': 0,
            'categories_updated': 0,
            'products_created': 0,
            'products_updated': 0,
            'modifiers_created': 0,
            'modifier_options_created': 0,
            'errors': []
        }
        
        with transaction.atomic():
            # ===== PROCESS CATEGORIES =====
            if 'Categories' in wb.sheetnames:
                ws_cat = wb['Categories']
                logger.info("Processing Categories sheet...")
                
                # Skip header row
                for row_idx, row in enumerate(ws_cat.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        company_code, brand_code, cat_name, parent_name, sort_order, icon, is_active = row[:7]
                        
                        if not company_code or not brand_code or not cat_name:
                            continue
                        
                        # Get brand with company validation
                        try:
                            brand = Brand.objects.select_related('company').get(
                                code=brand_code.strip(),
                                company__code=company_code.strip()
                            )
                        except Brand.DoesNotExist:
                            stats['errors'].append(
                                f"Row {row_idx}: Brand '{brand_code}' with Company '{company_code}' not found"
                            )
                            continue
                        
                        # Get parent category if specified
                        parent = None
                        if parent_name:
                            parent = Category.objects.filter(
                                brand=brand,
                                name=parent_name.strip()
                            ).first()
                            if not parent:
                                stats['errors'].append(
                                    f"Row {row_idx}: Parent category '{parent_name}' not found"
                                )
                        
                        # Create or update category
                        category, created = Category.objects.update_or_create(
                            brand=brand,
                            name=cat_name.strip(),
                            defaults={
                                'parent': parent,
                                'sort_order': int(sort_order) if sort_order else 0,
                                'icon': icon.strip() if icon else '',
                                'is_active': str(is_active).upper() == 'TRUE' if is_active else True
                            }
                        )
                        
                        if created:
                            stats['categories_created'] += 1
                        else:
                            stats['categories_updated'] += 1
                            
                    except Exception as e:
                        stats['errors'].append(f"Row {row_idx} (Categories): {str(e)}")
                        logger.error(f"Error processing category row {row_idx}: {traceback.format_exc()}")
            
            # ===== PROCESS PRODUCTS =====
            if 'Products' in wb.sheetnames:
                ws_prod = wb['Products']
                logger.info("Processing Products sheet...")
                
                # Define static image folder
                static_image_folder = Path(settings.BASE_DIR) / 'static' / 'product' / 'image'
                media_product_folder = Path(settings.MEDIA_ROOT) / 'product_photos'
                
                # Create folders if not exist
                static_image_folder.mkdir(parents=True, exist_ok=True)
                media_product_folder.mkdir(parents=True, exist_ok=True)
                
                for row_idx, row in enumerate(ws_prod.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        (company_code, brand_code, sku, prod_name, cat_name, description,
                         price, cost, printer_target, track_stock, stock_qty,
                         image_filename, sort_order, is_active) = row[:14]
                        
                        if not company_code or not brand_code or not sku or not prod_name:
                            continue
                        
                        # Get brand with company validation
                        try:
                            brand = Brand.objects.select_related('company').get(
                                code=brand_code.strip(),
                                company__code=company_code.strip()
                            )
                        except Brand.DoesNotExist:
                            stats['errors'].append(
                                f"Row {row_idx}: Brand '{brand_code}' with Company '{company_code}' not found"
                            )
                            continue
                        
                        # Get category
                        try:
                            category = Category.objects.get(brand=brand, name=cat_name.strip())
                        except Category.DoesNotExist:
                            stats['errors'].append(
                                f"Row {row_idx}: Category '{cat_name}' not found for brand '{brand_code}'"
                            )
                            continue
                        
                        # Create or update product
                        product, created = Product.objects.update_or_create(
                            brand=brand,
                            sku=sku.strip(),
                            defaults={
                                'name': prod_name.strip(),
                                'category': category,
                                'description': description.strip() if description else '',
                                'price': Decimal(str(price)) if price else Decimal('0'),
                                'cost': Decimal(str(cost)) if cost else Decimal('0'),
                                'printer_target': printer_target.strip() if printer_target else 'kitchen',
                                'track_stock': str(track_stock).upper() == 'TRUE' if track_stock else False,
                                'stock_quantity': Decimal(str(stock_qty)) if stock_qty else Decimal('0'),
                                'sort_order': int(sort_order) if sort_order else 0,
                                'is_active': str(is_active).upper() == 'TRUE' if is_active else True
                            }
                        )
                        
                        if created:
                            stats['products_created'] += 1
                        else:
                            stats['products_updated'] += 1
                        
                        # Process image if filename provided
                        if image_filename and str(image_filename).strip():
                            try:
                                filename = str(image_filename).strip()
                                source_path = static_image_folder / filename
                                
                                # Check if file exists in static/product/image
                                if not source_path.exists():
                                    stats['errors'].append(
                                        f"Row {row_idx}: Image file '{filename}' not found in static/product/image/"
                                    )
                                else:
                                    # Generate unique filename with timestamp
                                    file_ext = source_path.suffix
                                    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
                                    dest_path = media_product_folder / unique_filename
                                    
                                    # Copy file to media folder
                                    shutil.copy2(source_path, dest_path)
                                    
                                    # Delete existing photos for this product
                                    ProductPhoto.objects.filter(product=product).delete()
                                    
                                    # Create ProductPhoto with relative path
                                    ProductPhoto.objects.create(
                                        product=product,
                                        photo=f'product_photos/{unique_filename}',
                                        is_primary=True,
                                        sort_order=0
                                    )
                                    
                                    logger.info(f"Image copied for product {sku}: {filename} -> {unique_filename}")
                                    
                            except Exception as img_error:
                                stats['errors'].append(
                                    f"Row {row_idx}: Error processing image '{image_filename}': {str(img_error)}"
                                )
                                logger.error(f"Image error for row {row_idx}: {traceback.format_exc()}")
                            
                    except Exception as e:
                        stats['errors'].append(f"Row {row_idx} (Products): {str(e)}")
                        logger.error(f"Error processing product row {row_idx}: {traceback.format_exc()}")
            
            # ===== PROCESS MODIFIERS =====
            if 'Modifiers' in wb.sheetnames:
                ws_mod = wb['Modifiers']
                logger.info("Processing Modifiers sheet...")
                
                for row_idx, row in enumerate(ws_mod.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        (company_code, brand_code, mod_group, is_required, max_sel,
                         opt_name, price_adj, is_default, sort_order, is_active) = row[:10]
                        
                        if not company_code or not brand_code or not mod_group or not opt_name:
                            continue
                        
                        # Get brand with company validation
                        try:
                            brand = Brand.objects.select_related('company').get(
                                code=brand_code.strip(),
                                company__code=company_code.strip()
                            )
                        except Brand.DoesNotExist:
                            stats['errors'].append(
                                f"Row {row_idx}: Brand '{brand_code}' with Company '{company_code}' not found"
                            )
                            continue
                        
                        # Create or get modifier group
                        modifier, created = Modifier.objects.get_or_create(
                            brand=brand,
                            name=mod_group.strip(),
                            defaults={
                                'is_required': str(is_required).upper() == 'TRUE' if is_required else False,
                                'max_selections': int(max_sel) if max_sel else 1,
                                'is_active': True
                            }
                        )
                        
                        if created:
                            stats['modifiers_created'] += 1
                        
                        # Create modifier option
                        ModifierOption.objects.update_or_create(
                            modifier=modifier,
                            name=opt_name.strip(),
                            defaults={
                                'price_adjustment': Decimal(str(price_adj)) if price_adj else Decimal('0'),
                                'is_default': str(is_default).upper() == 'TRUE' if is_default else False,
                                'sort_order': int(sort_order) if sort_order else 0,
                                'is_active': str(is_active).upper() == 'TRUE' if is_active else True
                            }
                        )
                        stats['modifier_options_created'] += 1
                        
                    except Exception as e:
                        stats['errors'].append(f"Row {row_idx} (Modifiers): {str(e)}")
                        logger.error(f"Error processing modifier row {row_idx}: {traceback.format_exc()}")
        
        logger.info(f"Bulk import completed by {request.user.username}: {stats}")
        
        return JsonResponse({
            'success': True,
            'message': 'File uploaded and processed successfully',
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Bulk import error: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        }, status=500)


@login_required
def bulk_import_products_view(request):
    """Display bulk import products page (custom format)"""
    return render(request, 'settings/bulk_import_products.html')


@login_required
@require_http_methods(["POST"])
def upload_products_excel(request):
    """
    Upload and process custom products Excel format
    Columns: Company Code, Brand Code, Category, Menu Category, Nama Product, 
             PLU Product, Printer_Kitchen, Condiment Groups, Price Product, Image Product
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)
        
        excel_file = request.FILES['file']
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Please upload .xlsx or .xls file'
            }, status=400)
        
        # Load workbook
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active  # Assume single sheet
        
        # Get import options
        skip_duplicates = request.POST.get('skip_duplicates') == 'on'
        update_existing = request.POST.get('update_existing') == 'on'
        create_modifiers = request.POST.get('create_modifiers') == 'on'
        
        # Update option overrides skip option
        if update_existing:
            skip_duplicates = False
        
        stats = {
            'categories_created': 0,
            'products_created': 0,
            'products_updated': 0,
            'products_skipped': 0,
            'modifiers_created': 0,
            'modifier_options_created': 0,
            'product_modifiers_created': 0,
            'companies_created': 0,
            'brands_created': 0,
            'errors': [],
            'processed_rows': 0,
            'successful_rows': 0,
            'failed_rows': [],
            'detailed_log': []
        }
        
        # Define static image folder
        static_image_folder = Path(settings.BASE_DIR) / 'static' / 'product' / 'image'
        media_product_folder = Path(settings.MEDIA_ROOT) / 'product_photos'
        
        # Create folders if not exist
        static_image_folder.mkdir(parents=True, exist_ok=True)
        media_product_folder.mkdir(parents=True, exist_ok=True)
        
        # Cache for categories and modifiers
        category_cache = {}
        modifier_cache = {}
        
        with transaction.atomic():
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    (company_code, brand_code, category_name, menu_category_name, 
                     product_name, plu_product, printer_kitchen, condiment_groups,
                     price_product, image_product) = row[:10]
                    
                    if not company_code or not brand_code or not product_name or not plu_product:
                        continue
                    
                    # Get brand with company validation
                    try:
                        brand = Brand.objects.select_related('company').get(
                            code=brand_code.strip(),
                            company__code=company_code.strip()
                        )
                    except Brand.DoesNotExist:
                        stats['errors'].append(
                            f"Row {row_idx}: Brand '{brand_code}' with Company '{company_code}' not found"
                        )
                        continue
                    
                    # ===== PROCESS CATEGORIES =====
                    parent_category = None
                    final_category = None
                    
                    # Create/get parent category
                    if category_name:
                        cat_key = f"{brand.id}_{category_name.strip()}"
                        if cat_key not in category_cache:
                            parent_category, created = Category.objects.get_or_create(
                                brand=brand,
                                name=category_name.strip(),
                                parent=None,
                                defaults={'sort_order': 0, 'is_active': True}
                            )
                            category_cache[cat_key] = parent_category
                            if created:
                                stats['categories_created'] += 1
                        else:
                            parent_category = category_cache[cat_key]
                    
                    # Create/get menu category (child)
                    if menu_category_name:
                        menu_cat_key = f"{brand.id}_{menu_category_name.strip()}"
                        if menu_cat_key not in category_cache:
                            final_category, created = Category.objects.get_or_create(
                                brand=brand,
                                name=menu_category_name.strip(),
                                parent=parent_category,
                                defaults={'sort_order': 0, 'is_active': True}
                            )
                            category_cache[menu_cat_key] = final_category
                            if created:
                                stats['categories_created'] += 1
                        else:
                            final_category = category_cache[menu_cat_key]
                    else:
                        final_category = parent_category
                    
                    if not final_category:
                        stats['errors'].append(f"Row {row_idx}: No category specified")
                        continue
                    
                    # ===== PROCESS PRODUCT =====
                    sku = str(plu_product).strip()
                    
                    # Check for duplicate product
                    existing_product = None
                    if skip_duplicates or update_existing:
                        try:
                            existing_product = Product.objects.get(
                                brand=brand,
                                sku=sku,
                                name=product_name.strip(),
                                category=final_category
                            )
                        except Product.DoesNotExist:
                            existing_product = None
                    
                    # Skip if duplicate and not updating
                    if skip_duplicates and existing_product:
                        stats['products_skipped'] += 1
                        logger.info(f"Skipped duplicate product: {sku}")
                        continue
                    
                    # Map printer_kitchen
                    printer_map = {
                        'bar': 'bar',
                        'kitchen': 'kitchen',
                        'dessert': 'dessert'
                    }
                    printer_target = 'kitchen'
                    if printer_kitchen:
                        printer_target = printer_map.get(str(printer_kitchen).strip().lower(), 'kitchen')
                    
                    # Create or update product
                    if update_existing and existing_product:
                        # Update existing product
                        existing_product.name = product_name.strip()
                        existing_product.category = final_category
                        existing_product.description = ''
                        existing_product.price = Decimal(str(price_product)) if price_product else Decimal('0')
                        existing_product.printer_target = printer_target
                        existing_product.save()
                        
                        product = existing_product
                        stats['products_updated'] += 1
                        logger.info(f"Updated existing product: {sku}")
                    else:
                        # Create new product
                        product, created = Product.objects.get_or_create(
                            brand=brand,
                            sku=sku,
                            defaults={
                                'name': product_name.strip(),
                                'category': final_category,
                                'description': '',
                                'price': Decimal(str(price_product)) if price_product else Decimal('0'),
                                'cost': Decimal('0'),
                                'printer_target': printer_target,
                                'track_stock': False,
                                'stock_quantity': Decimal('0'),
                                'sort_order': 0,
                                'is_active': True
                            }
                        )
                        
                        if created:
                            stats['products_created'] += 1
                        else:
                            stats['products_updated'] += 1
                    
                    # ===== PROCESS IMAGE =====
                    if image_product and str(image_product).strip():
                        try:
                            image_path = str(image_product).strip()
                            # Extract filename from path (e.g., products/Menu_Image_20231203013316.jpg)
                            filename = Path(image_path).name
                            
                            source_path = static_image_folder / filename
                            
                            if source_path.exists():
                                # Generate unique filename
                                file_ext = source_path.suffix
                                unique_filename = f"{uuid.uuid4().hex}{file_ext}"
                                dest_path = media_product_folder / unique_filename
                                
                                # Copy file to media folder
                                shutil.copy2(source_path, dest_path)
                                
                                # Delete existing photos
                                ProductPhoto.objects.filter(product=product).delete()
                                
                                # Create ProductPhoto
                                ProductPhoto.objects.create(
                                    product=product,
                                    photo=f'product_photos/{unique_filename}',
                                    is_primary=True,
                                    sort_order=0
                                )
                            else:
                                stats['errors'].append(
                                    f"Row {row_idx}: Image '{filename}' not found in static/product/image/"
                                )
                        except Exception as img_error:
                            stats['errors'].append(
                                f"Row {row_idx}: Error processing image: {str(img_error)}"
                            )
                    
                    # ===== PROCESS MODIFIERS (Condiment Groups) =====
                    if create_modifiers and condiment_groups and str(condiment_groups).strip():
                        try:
                            groups_str = str(condiment_groups).strip()
                            if groups_str and groups_str not in ['', 'nan', 'None']:
                                # Split by comma
                                modifier_names = [m.strip() for m in groups_str.split(',') if m.strip()]
                                
                                for mod_name in modifier_names:
                                    mod_key = f"{brand.id}_{mod_name}"
                                    
                                    # Get or create modifier
                                    if mod_key not in modifier_cache:
                                        modifier, created = Modifier.objects.get_or_create(
                                            brand=brand,
                                            name=mod_name,
                                            defaults={
                                                'is_required': False,
                                                'max_selections': 1,
                                                'is_active': True
                                            }
                                        )
                                        modifier_cache[mod_key] = modifier
                                        if created:
                                            stats['modifiers_created'] += 1
                                    else:
                                        modifier = modifier_cache[mod_key]
                                    
                                    # Link product to modifier via ProductModifier
                                    product_modifier, created = ProductModifier.objects.get_or_create(
                                        product=product,
                                        modifier=modifier,
                                        defaults={'sort_order': len(product.product_modifiers.all())}
                                    )
                                    if created:
                                        stats['product_modifiers_created'] += 1
                                    
                        except Exception as mod_error:
                            stats['errors'].append(
                                f"Row {row_idx}: Error processing modifiers: {str(mod_error)}"
                            )
                    
                except Exception as e:
                    stats['errors'].append(f"Row {row_idx}: {str(e)}")
                    logger.error(f"Error processing row {row_idx}: {traceback.format_exc()}")
        
        logger.info(f"Products bulk import completed by {request.user.username}: {stats}")
        
        return JsonResponse({
            'success': True,
            'message': 'Products imported successfully',
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Products bulk import error: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        }, status=500)
